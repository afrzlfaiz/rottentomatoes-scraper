import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Any
from urllib.parse import urlencode, urlparse
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from colorama import init, Fore, Style, Back
from datetime import datetime
import time

# Initialize colorama
init(autoreset=True)


MOVIE_REVIEWS_URL = "https://www.rottentomatoes.com/napi/rtcf/v1/movies/{movie_id}/reviews"
TV_SEASON_REVIEWS_URL = "https://www.rottentomatoes.com/napi/rtcf/v1/tv/seasons/{movie_id}/reviews"
PROPS_PATTERN = re.compile(
    r'<script\s+data-json="props"\s+type="application/json">\s*(\{.*?\})\s*</script>',
    re.DOTALL,
)
CLI_HEADER = r"""$$$$$$$\             $$\     $$\                          $$$$$$\  $$\       $$$$$$\
$$  __$$\            $$ |    $$ |                        $$  __$$\ $$ |      \_$$  _|
$$ |  $$ | $$$$$$\ $$$$$$\ $$$$$$\    $$$$$$\  $$$$$$$\  $$ /  \__|$$ |        $$ |
$$$$$$$  |$$  __$$\\_$$  _|\_$$  _|  $$  __$$\ $$  __$$\ $$ |      $$ |        $$ |
$$  __$$< $$ /  $$ | $$ |    $$ |    $$$$$$$$ |$$ |  $$ |$$ |      $$ |        $$ |
$$ |  $$ |$$ |  $$ | $$ |$$\ $$ |$$\ $$   ____|$$ |  $$ |$$ |  $$\ $$ |        $$ |
$$ |  $$ |\$$$$$$  | \$$$$  |\$$$$  |\$$$$$$$\ $$ |  $$ |\$$$$$$  |$$$$$$$$\ $$$$$$\
\__|  \__| \______/   \____/  \____/  \_______|\__|  \__| \______/ \________|\______|  """

# Color schemes
class Colors:
    HEADER = Fore.MAGENTA + Style.BRIGHT
    SUCCESS = Fore.GREEN + Style.BRIGHT
    WARNING = Fore.YELLOW + Style.BRIGHT
    ERROR = Fore.RED + Style.BRIGHT
    INFO = Fore.CYAN + Style.BRIGHT
    DEBUG = Fore.BLUE + Style.BRIGHT
    PROCESS = Fore.WHITE + Back.BLUE
    HIGHLIGHT = Fore.YELLOW + Back.BLACK
    RESET_ALL = Style.RESET_ALL

# Logger class for better output formatting
class Logger:
    def __init__(self):
        self.start_time = datetime.now()

    def log(self, level: str, message: str, emoji: str = ""):
        timestamp = datetime.now().strftime("%H:%M:%S")
        level_colors = {
            "INFO": Colors.INFO,
            "SUCCESS": Colors.SUCCESS,
            "WARNING": Colors.WARNING,
            "ERROR": Colors.ERROR,
            "PROCESS": Colors.PROCESS,
            "DEBUG": Colors.DEBUG,
            "HEADER": Colors.HEADER,
        }

        color = level_colors.get(level, Colors.INFO)
        emoji_prefix = f" {emoji}" if emoji else ""

        print(f"{color}[{timestamp}] {level}{emoji_prefix}: {message}{Style.RESET_ALL}")

    def log_progress(self, current: int, total: int, suffix: str = ""):
        percent = current / total * 100
        bar_length = 30
        filled_length = int(bar_length * current // total)
        bar = '█' * filled_length + '-' * (bar_length - filled_length)

        print(f"{Colors.PROCESS}\r[{bar}] {current}/{total} ({percent:.1f}%) {suffix}{Style.RESET_ALL}", end='\r')
        if current == total:
            print()  # New line when complete

    def log_with_spinner(self, message: str, duration: float = 0.5):
        spinner = ['⠋', '⠙', '⠹', '⠸', '⠼', '⠴', '⠦', '⠧', '⠇', '⠏']
        i = 0
        start = time.time()

        while time.time() - start < duration:
            print(f"{Colors.INFO}\r{message} {spinner[i % len(spinner)]}{Style.RESET_ALL}", end='\r')
            time.sleep(0.1)
            i += 1
        print()  # New line

# Initialize logger
logger = Logger()


def star_to_number(value: str | None) -> float | None:
    if not value or not value.startswith("STAR_"):
        return None
    return float(value.removeprefix("STAR_").replace("_", "."))


def fetch_page(
    movie_id: str,
    content_type: str,
    review_type: str,
    after: str = "",
    page_count: int = 20,
    verified: bool = False,
    top_only: bool = False,
) -> dict[str, Any]:
    params = {
        "after": after,
        "before": "",
        "pageCount": page_count,
        "topOnly": str(top_only).lower(),
        "type": review_type,
        "verified": str(verified).lower(),
    }
    if content_type == "tv":
        url = f"{TV_SEASON_REVIEWS_URL.format(movie_id=movie_id)}?{urlencode(params)}"
    else:
        url = f"{MOVIE_REVIEWS_URL.format(movie_id=movie_id)}?{urlencode(params)}"

    request = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json",
        },
    )

    try:
        with urlopen(request) as response:
            data = json.load(response)
            if isinstance(data, list):
                data = {
                    "reviews": [item for item in data if isinstance(item, dict)],
                    "pageInfo": {"hasNextPage": False, "endCursor": ""},
                }
            return data
    except HTTPError as e:
        logger.log("ERROR", f"HTTP Error {e.code}: {e.reason}", "❌")
        raise
    except Exception as e:
        logger.log("ERROR", f"Request failed: {str(e)}", "❌")
        raise


def fetch_text(url: str) -> str:
    request = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html,application/xhtml+xml",
        },
    )

    with urlopen(request) as response:
        return response.read().decode("utf-8", errors="replace")


def normalize_rt_url(rt_url: str) -> str:
    parsed = urlparse(rt_url.strip())
    if not parsed.scheme or not parsed.netloc:
        raise ValueError("Please enter a valid Rotten Tomatoes URL.")
    if parsed.netloc not in {"www.rottentomatoes.com", "rottentomatoes.com"}:
        raise ValueError("URL must be from rottentomatoes.com")

    path = parsed.path.rstrip("/")
    if not path or path == "/":
        raise ValueError("Invalid Rotten Tomatoes URL path.")

    return f"https://www.rottentomatoes.com{path}"


def is_tv_url(rt_url: str) -> bool:
    path = urlparse(normalize_rt_url(rt_url)).path.rstrip("/")
    return path.startswith("/tv/")


def tv_url_has_season(rt_url: str) -> bool:
    path = urlparse(normalize_rt_url(rt_url)).path.rstrip("/")
    return bool(re.search(r"/s\d{1,2}$", path))


def append_tv_season(rt_url: str, season: int) -> str:
    base_url = normalize_rt_url(rt_url)
    if season <= 0:
        raise ValueError("Season number must be greater than 0.")
    return f"{base_url}/s{season:02d}"


def ensure_tv_season_url(rt_url: str, interactive: bool) -> str:
    normalized = normalize_rt_url(rt_url)
    if not is_tv_url(normalized) or tv_url_has_season(normalized):
        return normalized

    if not interactive:
        raise ValueError("TV show URL must include season path (example: /s01).")

    season = prompt_positive_int("TV show detected. Which season number? ")
    season_url = append_tv_season(normalized, season)
    logger.log("INFO", f"Using TV season URL: {season_url}", "📺")
    return season_url


def resolve_movie_id_from_url(rt_url: str) -> tuple[str, str, str | None]:
    base_url = normalize_rt_url(rt_url)
    reviews_url = f"{base_url}/reviews/all-audience"

    logger.log("INFO", f"Accessing URL: {reviews_url}")
    logger.log_with_spinner("Fetching web page", 1.0)

    html = fetch_text(reviews_url)

    match = PROPS_PATTERN.search(html)
    if not match:
        logger.log("ERROR", "Failed to find JSON props on audience review page")
        raise ValueError("Failed to find JSON props on audience review page.")

    props = json.loads(match.group(1))
    media = props.get("media") or {}
    movie_id = media.get("emsId")
    title = media.get("title")
    if not movie_id:
        logger.log("ERROR", "emsId not found on audience review page")
        raise ValueError("emsId not found on audience review page.")

    logger.log("SUCCESS", f"Content ID found: {movie_id}", "🎯")
    if title:
        logger.log("SUCCESS", f"Title found: {title}", "🎬")
    return str(movie_id), reviews_url, title


def slug_from_url(rt_url: str) -> str:
    path = urlparse(normalize_rt_url(rt_url)).path.rstrip("/")
    return path.rsplit("/", 1)[-1]


def normalize_review(review: dict[str, Any], review_type: str) -> dict[str, Any]:
    if review_type == "audience":
        return {
            "id": review.get("ratingId"),
            "type": review_type,
            "rating": review.get("rating"),
            "rating_value": star_to_number(review.get("rating")),
            "verified": review.get("isVerified"),
            "display_name": review.get("displayName"),
            "review": review.get("review"),
            "created_at": review.get("createDate"),
        }

    critic = review.get("critic") or {}
    publication = review.get("publication") or {}
    return {
        "id": review.get("reviewId"),
        "type": review_type,
        "score_sentiment": review.get("scoreSentiment"),
        "original_score": review.get("originalScore"),
        "is_top_review": review.get("isTopReview"),
        "critic_name": critic.get("displayName"),
        "publication_name": publication.get("name"),
        "review_quote": review.get("reviewQuote"),
        "created_at": review.get("createDate"),
        "review_url": review.get("publicationReviewUrl"),
    }


def fetch_reviews(
    movie_id: str,
    content_type: str,
    review_type: str,
    max_reviews: int,
    verified: bool,
    top_only: bool,
) -> list[dict[str, Any]]:
    after = ""
    reviews: list[dict[str, Any]] = []

    logger.log("INFO", f"Starting to fetch {max_reviews} {review_type} reviews...", "📥")
    logger.log_with_spinner(f"Fetching {review_type} reviews", 1.0)

    page_num = 1
    while len(reviews) < max_reviews:
        page_count = min(20, max_reviews - len(reviews))

        # Show progress bar
        logger.log_progress(len(reviews), max_reviews, f"Review {review_type}")

        payload = fetch_page(
            movie_id=movie_id,
            content_type=content_type,
            review_type=review_type,
            after=after,
            page_count=page_count,
            verified=verified,
            top_only=top_only,
        )

        page_reviews = payload.get("reviews") or []
        reviews.extend(normalize_review(review, review_type) for review in page_reviews)

        if len(reviews) >= max_reviews:
            break

        page_info = payload.get("pageInfo") or {}
        if not page_info.get("hasNextPage"):
            logger.log("WARNING", "No next page")
            break

        after = page_info.get("endCursor") or ""
        if not after:
            logger.log("WARNING", "Empty cursor, stopping")
            break

        page_num += 1
        time.sleep(0.5)  # Small delay to avoid rate limiting

    logger.log("SUCCESS", f"Completed: {len(reviews)}/{max_reviews} {review_type} reviews fetched", "✅")
    return reviews[:max_reviews]


def clear_screen() -> None:
    os.system('cls' if os.name == 'nt' else 'clear')


def render_header() -> None:
    clear_screen()
    print(Colors.HEADER + CLI_HEADER + Style.RESET_ALL)
    print(Colors.HIGHLIGHT + "=" * 80 + Style.RESET_ALL)
    print()


def prompt_non_empty(message: str) -> str:
    error_message = ""
    while True:
        render_header()
        if error_message:
            logger.log("ERROR", error_message, "❌")
            print()
        value = input(f"{Colors.INFO}{message}{Style.RESET_ALL}").strip()
        if value:
            return value
        error_message = "Input cannot be empty."
        time.sleep(0.1)  # Small delay to prevent rapid rendering


def prompt_choice(message: str, options: dict[str, str]) -> str:
    error_message = ""
    while True:
        render_header()
        if error_message:
            logger.log("ERROR", error_message, "❌")
            print()

        print(f"{Colors.INFO}{message}{Style.RESET_ALL}")
        for key, label in options.items():
            print(f"{Colors.HIGHLIGHT}{key}. {Style.RESET_ALL}{label}")
        choice = input(f"{Colors.INFO}Choose a number: {Style.RESET_ALL}").strip()
        if choice in options:
            return choice
        error_message = "Invalid choice."


def prompt_positive_int(message: str) -> int:
    error_message = ""
    while True:
        render_header()
        if error_message:
            logger.log("ERROR", error_message, "❌")
            print()

        value = input(f"{Colors.INFO}{message}{Style.RESET_ALL}").strip()
        try:
            number = int(value)
        except ValueError:
            error_message = "Enter an integer."
            continue
        if number > 0:
            logger.log("SUCCESS", f"Input: {number}", "✓")
            return number
        error_message = "Number must be greater than 0."


def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def write_json(output_path: Path, payload: dict[str, Any]) -> None:
    logger.log("INFO", f"Writing JSON to {output_path}...", "💾")
    ensure_parent_dir(output_path)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.log("SUCCESS", "JSON file written successfully", "✅")


def write_csv(base_path: Path, payload: dict[str, Any], append_mode: bool = False) -> list[Path]:
    written_files: list[Path] = []
    base_stem = base_path.stem
    suffix = base_path.suffix if base_path.suffix.lower() == ".csv" else ".csv"

    for review_type, rows in payload["reviews"].items():
        logger.log("INFO", f"Writing CSV for {review_type} reviews...", "📊")
        output_path = base_path.with_name(f"{base_stem}_{review_type}{suffix}")
        ensure_parent_dir(output_path)

        enriched_rows = [
            {
                "title": payload.get("title"),
                "movie_id": payload.get("movie_id"),
                **row,
            }
            for row in rows
        ]

        fieldnames = list(enriched_rows[0].keys()) if enriched_rows else ["title", "movie_id", "type"]

        if append_mode and output_path.exists():
            with output_path.open("a", newline="", encoding="utf-8-sig") as file:
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writerows(enriched_rows)
        else:
            with output_path.open("w", newline="", encoding="utf-8-sig") as file:
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(enriched_rows)
        written_files.append(output_path)
        logger.log("SUCCESS", f"CSV {'appended to' if append_mode else 'written'}: {output_path}", "✅")

    return written_files


def write_excel(output_path: Path, payload: dict[str, Any], append_mode: bool = False) -> None:
    logger.log("INFO", f"Writing Excel to {output_path}...", "📈")
    ensure_parent_dir(output_path)

    if append_mode and output_path.exists():
        workbook = load_workbook(output_path)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    for review_type, rows in payload["reviews"].items():
        logger.log("INFO", f"Creating sheet for {review_type} reviews...", "📊")
        sheet_name = review_type[:31]

        enriched_rows = [
            {
                "title": payload.get("title"),
                "movie_id": payload.get("movie_id"),
                **row,
            }
            for row in rows
        ]

        if not enriched_rows:
            continue

        headers = list(enriched_rows[0].keys())

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.append(headers)

        for row in enriched_rows:
            sheet.append([row.get(header) for header in headers])

    workbook.save(output_path)
    logger.log("SUCCESS", "Excel file written successfully", "✅")


def export_reviews(output_base: str, output_format: str, payload: dict[str, Any], append_mode: bool = False) -> list[Path]:
    base_path = Path(output_base)
    written_files: list[Path] = []

    if output_format in {"json", "all"}:
        logger.log("INFO", "Exporting JSON files...", "📄")
        json_base = base_path.with_suffix("") if base_path.suffix.lower() == ".json" else base_path
        for review_type, rows in payload["reviews"].items():
            json_path = json_base.with_name(f"{json_base.name}_{review_type}").with_suffix(".json")
            json_payload = {
                **payload,
                "review_types": [review_type],
                "reviews": {review_type: rows},
            }
            write_json(json_path, json_payload)
            written_files.append(json_path)
            logger.log("SUCCESS", f"JSON file saved: {json_path}", "✅")

    if output_format in {"csv", "all"}:
        logger.log("INFO", "Exporting CSV files...", "📊")
        csv_before = len(written_files)
        written_files.extend(write_csv(base_path, payload, append_mode=append_mode))
        logger.log("SUCCESS", f"CSV files saved: {len(written_files) - csv_before} files", "✅")

    if output_format in {"excel", "all"}:
        logger.log("INFO", "Exporting Excel files...", "📈")
        excel_base = base_path.with_suffix("") if base_path.suffix.lower() == ".xlsx" else base_path
        for review_type, rows in payload["reviews"].items():
            excel_path = excel_base.with_name(f"{excel_base.name}_{review_type}").with_suffix(".xlsx")
            excel_payload = {
                **payload,
                "review_types": [review_type],
                "reviews": {review_type: rows},
            }
            write_excel(excel_path, excel_payload, append_mode=append_mode)
            written_files.append(excel_path)
            logger.log("SUCCESS", f"Excel file saved: {excel_path}", "✅")

    return written_files


def prompt_interactive_batch() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    targets: list[dict[str, Any]] = []

    def _resolve_url(prompt_text: str) -> tuple[str, str, str | None]:
        source_url = prompt_non_empty(prompt_text)
        source_url = ensure_tv_season_url(source_url, interactive=True)
        movie_id, _, title = resolve_movie_id_from_url(source_url)
        logger.log("SUCCESS", f"Content ID: {movie_id}", "🎯")
        if title:
            logger.log("SUCCESS", f"Title: {title}", "🎬")
        return movie_id, title, source_url

    movie_id, title, source_url = _resolve_url("Enter Rotten Tomatoes movie URL: ")

    review_type_choice = prompt_choice(
        "Select review type:",
        {"1": "audience", "2": "critic", "3": "both"},
    )
    output_format_choice = prompt_choice(
        "Select output format:",
        {"1": "json", "2": "csv", "3": "excel", "4": "all"},
    )
    count = prompt_positive_int("Number of reviews to fetch: ")
    output_base = prompt_non_empty("Base output file name/path (e.g., output/review): ")

    review_type_map = {"1": "audience", "2": "critic", "3": "both"}
    output_format_map = {"1": "json", "2": "csv", "3": "excel", "4": "all"}

    targets.append({
        "movie_id": movie_id,
        "title": title,
        "url": source_url,
        "count": count,
    })

    while True:
        add_another = prompt_choice(
            "Add another film/TV show?",
            {"1": "Yes, add another", "2": "No, start scraping"},
        )
        if add_another == "2":
            break

        movie_id, title, source_url = _resolve_url("Enter Rotten Tomatoes URL for next film/TV show: ")
        count = prompt_positive_int("Number of reviews to fetch: ")

        targets.append({
            "movie_id": movie_id,
            "title": title,
            "url": source_url,
            "count": count,
        })

    common_settings = {
        "type": review_type_map[review_type_choice],
        "format": output_format_map[output_format_choice],
        "output": output_base,
    }
    return targets, common_settings


def show_batch_summary(targets: list[dict[str, Any]], settings: dict[str, Any]) -> None:
    render_header()
    logger.log("HEADER", "Batch Configuration Summary:", "📋")
    print(f"{Colors.INFO}• Review type: {Style.RESET_ALL}{settings['type']}")
    print(f"{Colors.INFO}• Output format: {Style.RESET_ALL}{settings['format']}")
    print(f"{Colors.INFO}• Output file base: {Style.RESET_ALL}{settings['output']}")
    print()
    for i, t in enumerate(targets, 1):
        print(f"{Colors.HIGHLIGHT}  #{i} {Style.RESET_ALL}{t['title'] or '-'}  ({t['count']} reviews)")
        print(f"      {Colors.INFO}{t['url']}{Style.RESET_ALL}")
    print()
    logger.log("INFO", "Press Enter to start scraping...", "🚀")
    input()


def _scrape_one_film(
    target: dict[str, Any],
    common_settings: dict[str, Any],
    args: argparse.Namespace,
) -> dict[str, Any]:
    movie_id = target.get("movie_id")
    source_url = target.get("url")
    title = target.get("title")
    count = target.get("count")
    review_type_arg = common_settings["type"]
    interactive = common_settings.get("interactive", False)

    if source_url:
        source_url = ensure_tv_season_url(source_url, interactive=interactive)
        content_type = "tv" if is_tv_url(source_url) else "movie"
    else:
        content_type = "movie"

    if not movie_id and source_url:
        logger.log("INFO", "Fetching movie ID from URL...", "🔍")
        movie_id, reviews_url, title = resolve_movie_id_from_url(source_url)
        logger.log("SUCCESS", f"Content ID: {movie_id}", "🎯")
        if title:
            logger.log("SUCCESS", f"Title: {title}", "🎬")
    elif not movie_id:
        raise ValueError("Neither movie_id nor URL provided for target")
    else:
        reviews_url = None
        logger.log("INFO", f"Processing: {title or movie_id}", "🎬")

    review_types = [review_type_arg] if review_type_arg != "both" else ["audience", "critic"]
    reviews = {
        review_type: fetch_reviews(
            movie_id=movie_id,
            content_type=content_type,
            review_type=review_type,
            max_reviews=count,
            verified=args.verified if review_type == "audience" else False,
            top_only=args.top_only,
        )
        for review_type in review_types
    }

    result = {
        "movie_id": movie_id,
        "title": title,
        "movie_url": source_url,
        "source_reviews_url": reviews_url,
        "requested_count": count,
        "review_types": review_types,
        "reviews": reviews,
    }

    total_reviews = sum(len(reviews[rt]) for rt in reviews)
    logger.log("SUCCESS", f"{title or movie_id}: {total_reviews} reviews fetched", "✅")
    return result


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    render_header()
    logger.log("HEADER", "🍅 Rotten Tomatoes Scraper - Modern Version", "🎬")
    logger.log("INFO", "Script to scrape reviews from Rotten Tomatoes", "📊")
    print(Colors.HIGHLIGHT + "=" * 80 + Style.RESET_ALL)
    time.sleep(1)

    parser = argparse.ArgumentParser(
        description="Scrape Rotten Tomatoes reviews via public JSON endpoint."
    )
    parser.add_argument(
        "movie_id",
        nargs="?",
        help="Rotten Tomatoes Movie ID. If empty, script will prompt for movie URL interactively.",
    )
    parser.add_argument("--url", help="Rotten Tomatoes movie URL, example: https://www.rottentomatoes.com/m/lee_cronins_the_mummy")
    parser.add_argument(
        "--batch-file",
        help="Text file with one Rotten Tomatoes URL per line for batch scraping (CLI mode only)",
    )
    parser.add_argument(
        "--type",
        choices=["audience", "critic", "both"],
        default="both",
        help="Type of reviews to fetch",
    )
    parser.add_argument("--count", type=int, default=20, help="Number of reviews to fetch per film/type")
    parser.add_argument("--verified", action="store_true", help="Audience only: fetch verified reviews only")
    parser.add_argument("--top-only", action="store_true", help="Fetch top reviews only if available")
    parser.add_argument("--output", help="Base output file path. If empty in interactive mode, will be prompted")
    parser.add_argument(
        "--format",
        choices=["json", "csv", "excel", "all"],
        default="json",
        help="Output file format",
    )
    args = parser.parse_args()

    interactive_mode = not args.movie_id and not args.url and not args.batch_file

    if interactive_mode:
        logger.log("INFO", "Interactive mode enabled", "🔄")
        targets, common_settings = prompt_interactive_batch()
        show_batch_summary(targets, common_settings)
        common_settings["interactive"] = True
    else:
        targets: list[dict[str, Any]] = []
        if args.batch_file:
            batch_path = Path(args.batch_file)
            if not batch_path.exists():
                logger.log("ERROR", f"Batch file not found: {args.batch_file}", "❌")
                sys.exit(1)
            urls = [line.strip() for line in batch_path.read_text(encoding="utf-8").splitlines() if line.strip()]
            logger.log("INFO", f"Loaded {len(urls)} URLs from batch file", "📄")
            for url in urls:
                targets.append({"url": url, "count": args.count})
        elif args.url:
            targets.append({"url": args.url, "count": args.count})
        elif args.movie_id:
            targets.append({"movie_id": args.movie_id, "count": args.count})

        if not args.output:
            logger.log("ERROR", "--output is required in CLI mode", "❌")
            sys.exit(1)

        common_settings = {
            "type": args.type,
            "format": args.format,
            "output": args.output,
            "interactive": False,
        }

    logger.log("INFO", f"Review type: {common_settings['type']}", "📝")
    logger.log("INFO", f"Output format: {common_settings['format']}", "💾")
    logger.log("INFO", f"Total films to process: {len(targets)}", "🎬")

    all_written_files: list[Path] = []
    grand_total_reviews = 0

    for i, target in enumerate(targets):
        render_header()
        logger.log("HEADER", f"Processing [{i+1}/{len(targets)}]", "🎬")
        append_mode = (i > 0)
        output_base = common_settings.get("output")
        output_format = common_settings.get("format", "json")

        try:
            result = _scrape_one_film(target, common_settings, args)
            grand_total_reviews += sum(len(result["reviews"][rt]) for rt in result["reviews"])

            if output_base:
                logger.log("INFO", "Saving results to file...", "💾")
                logger.log_with_spinner("Exporting data", 1.0)
                written_files = export_reviews(output_base, output_format, result, append_mode=append_mode)
                if i == 0:
                    all_written_files = written_files
        except HTTPError as error:
            if error.code == 404:
                logger.log("ERROR", "Film/TV show has no rating on Rotten Tomatoes", "❌")
            else:
                logger.log("ERROR", f"HTTP Error {error.code}: {error.reason}", "❌")
            continue
        except Exception as error:
            logger.log("ERROR", f"Error: {str(error)}", "❌")
            continue

    clear_screen()
    print(Colors.SUCCESS + COMPLETION_ART + Style.RESET_ALL)
    logger.log("HEADER", "✅ Scraping completed!", "🎬")
    logger.log("SUCCESS", f"Total films: {len(targets)}", "🎬")
    logger.log("SUCCESS", f"Total reviews: {grand_total_reviews}", "📊")
    if all_written_files:
        logger.log("SUCCESS", "Output files:", "📁")
        for file_path in all_written_files:
            print(f"  {Colors.SUCCESS}✓ {file_path}{Style.RESET_ALL}")
    elif not output_base:
        logger.log("INFO", "No output file specified, results not saved to disk", "ℹ️")
    logger.log("SUCCESS", "Thank you for using the Rotten Tomatoes Scraper!", "🍅")


# ASCII art for completion
COMPLETION_ART = """
    🎬     🍅     ✨
   ( ●  ● )   ( ●  ● )  ( ●  ● )
  --------  --------  --------
   ROTTEN   TOMATOES   SCRAPER
   --------  --------  --------
        🎉   COMPLETED!   🎉
"""

if __name__ == "__main__":
    main()
